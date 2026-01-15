"""
SONGXS - Music Metadata Extractor
VERSIÓN 2: Con sistema de usuarios y autenticación
"""

from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, flash
import spotipy
from spotipy.oauth2 import SpotifyClientCredentials
import pandas as pd
import os
from datetime import datetime
import io
# Hola git
# ============================================
# NUEVO: Importar sistema de autenticación
# ============================================
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_bcrypt import Bcrypt

# Importar modelos actualizados
from models import (
    db, init_db, 
    add_track, search_tracks, get_all_tracks, get_catalog_stats,
    User, Track, ArtistCache, SplitSheet, Collaborator,
    create_user, get_user_by_email, verify_password,
    get_all_users, get_user_catalog_admin, get_global_stats
)

app = Flask(__name__)

# ============================================
# CONFIGURACIÓN
# ============================================

# Spotify API
SPOTIFY_CLIENT_ID = 'f414ec6cf4aa4cab9c46e3be6976ebf1'
SPOTIFY_CLIENT_SECRET = '39b3316f2faf4ac8898806b85a234db7'

# NUEVO: Last.fm API
LASTFM_API_KEY = '3af3d454901ce2b28e3d47d9f388d087'

# NUEVO: RapidAPI (Instagram + TikTok) - DESACTIVADO (sin créditos)
# Si tienes créditos, pon tu key aquí. Si no, déjalo vacío.
RAPIDAPI_KEY = ''  # Deja vacío si no tienes créditos
INSTAGRAM_API_HOST = 'instagram120.p.rapidapi.com'
TIKTOK_API_HOST = 'tiktok-api23.p.rapidapi.com'

# Base de datos
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///songxs.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# NUEVO: Secret key para sesiones
app.config['SECRET_KEY'] = 'tu-secret-key-super-secreta-cambiala-en-produccion'
# Esta key se usa para encriptar las cookies de sesión

# Inicializar extensiones
init_db(app)
bcrypt = Bcrypt(app)

# ============================================
# NUEVO: Configurar Flask-Login
# ============================================
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'  # Página de login
login_manager.login_message = 'Please log in to access this page.'

@login_manager.user_loader
def load_user(user_id):
    """
    Flask-Login usa esto para cargar el usuario actual
    Se ejecuta en cada request
    """
    return User.query.get(int(user_id))

# Inicializar Spotify
auth_manager = SpotifyClientCredentials(
    client_id=SPOTIFY_CLIENT_ID,
    client_secret=SPOTIFY_CLIENT_SECRET
)
sp = spotipy.Spotify(auth_manager=auth_manager)


# ============================================
# FUNCIONES HELPER: APIs Extendidas
# ============================================

def get_audio_features(track_id):
    """
    Obtiene audio features de Spotify (tempo, energy, etc.)
    
    Args:
        track_id (str): ID del track en Spotify
    
    Returns:
        dict: Audio features o None si hay error
    """
    try:
        features = sp.audio_features([track_id])[0]
        if features:
            return {
                'tempo': features.get('tempo'),
                'energy': features.get('energy'),
                'danceability': features.get('danceability'),
                'valence': features.get('valence'),
                'acousticness': features.get('acousticness'),
                'instrumentalness': features.get('instrumentalness'),
                'key': features.get('key'),
                'mode': features.get('mode'),
                'time_signature': features.get('time_signature')
            }
    except:
        return None


def get_artist_stats(artist_id):
    """
    Obtiene stats del artista de Spotify
    
    Args:
        artist_id (str): ID del artista en Spotify
    
    Returns:
        dict: Stats del artista o None si hay error
    """
    try:
        artist = sp.artist(artist_id)
        
        # Obtener la mejor imagen disponible
        artist_image = None
        if artist.get('images') and len(artist['images']) > 0:
            # Preferir imagen mediana (index 1) o la primera disponible
            artist_image = artist['images'][1]['url'] if len(artist['images']) > 1 else artist['images'][0]['url']
        
        return {
            'artist_id': artist_id,
            'followers': artist['followers']['total'],
            'popularity': artist['popularity'],
            'genres': ', '.join(artist['genres']) if artist['genres'] else '',
            'image_url': artist_image
        }
    except:
        return None


def get_lastfm_data(track_name, artist_name):
    """
    Obtiene datos de Last.fm (tags, play count, etc.)
    
    Args:
        track_name (str): Nombre del track
        artist_name (str): Nombre del artista
    
    Returns:
        dict: Datos de Last.fm o None si hay error
    """
    import requests
    
    try:
        # Get track info
        url = 'http://ws.audioscrobbler.com/2.0/'
        params = {
            'method': 'track.getInfo',
            'api_key': LASTFM_API_KEY,
            'artist': artist_name,
            'track': track_name,
            'format': 'json'
        }
        
        response = requests.get(url, params=params, timeout=5)
        
        if response.status_code == 200:
            data = response.json()
            
            if 'track' in data:
                track_info = data['track']
                
                # Extract tags
                tags = []
                if 'toptags' in track_info and 'tag' in track_info['toptags']:
                    tags = [tag['name'] for tag in track_info['toptags']['tag'][:5]]  # Top 5 tags
                
                return {
                    'tags': ', '.join(tags) if tags else '',
                    'playcount': int(track_info.get('playcount', 0)),
                    'listeners': int(track_info.get('listeners', 0))
                }
    except:
        pass
    
    return None


def get_artist_location(artist_name):
    """
    Obtiene la ubicación del artista de MusicBrainz
    
    Args:
        artist_name (str): Nombre del artista
    
    Returns:
        dict: Ubicación del artista o None si hay error
    """
    import requests
    import time
    
    try:
        # MusicBrainz API
        # IMPORTANTE: Respetar rate limit (1 request por segundo)
        url = 'https://musicbrainz.org/ws/2/artist/'
        
        # Buscar artista
        params = {
            'query': f'artist:{artist_name}',
            'fmt': 'json',
            'limit': 1
        }
        
        headers = {
            'User-Agent': 'SONGXS/1.0 (contact@songxs.com)'  # MusicBrainz requiere User-Agent
        }
        
        response = requests.get(url, params=params, headers=headers, timeout=5)
        
        if response.status_code == 200:
            data = response.json()
            
            if 'artists' in data and len(data['artists']) > 0:
                artist = data['artists'][0]
                
                # Extraer ubicación
                country = None
                city = None
                area_name = None
                
                # País
                if 'country' in artist:
                    country = artist['country']
                
                # Área (puede ser ciudad o país)
                if 'area' in artist and artist['area']:
                    area_name = artist['area'].get('name')
                    
                    # Si tiene "begin-area" es la ciudad de nacimiento
                    if 'begin-area' in artist and artist['begin-area']:
                        city = artist['begin-area'].get('name')
                
                # Si no encontramos ciudad pero tenemos área, usar área como ciudad
                if not city and area_name and country != area_name:
                    city = area_name
                
                return {
                    'country': country,
                    'city': city
                }
        
        # Rate limiting: esperar 1 segundo
        time.sleep(1)
        
    except Exception as e:
        print(f"⚠️  Error obteniendo ubicación de {artist_name}: {str(e)}")
    
    return None


def get_artist_location(artist_name):
    """
    Obtiene ubicación del artista desde MusicBrainz
    
    Args:
        artist_name (str): Nombre del artista
    
    Returns:
        dict: {'country': 'Puerto Rico', 'city': 'San Juan'} o None
    """
    import requests
    import time
    
    try:
        # MusicBrainz API
        url = 'https://musicbrainz.org/ws/2/artist/'
        
        # Headers requeridos (User-Agent es obligatorio en MusicBrainz)
        headers = {
            'User-Agent': 'SONGXS/1.0 (contact@songxs.com)',
            'Accept': 'application/json'
        }
        
        # Buscar artista por nombre
        params = {
            'query': f'artist:"{artist_name}"',
            'fmt': 'json',
            'limit': 1
        }
        
        response = requests.get(url, headers=headers, params=params, timeout=10)
        
        # MusicBrainz requiere rate limiting (1 request/second)
        time.sleep(1.1)
        
        if response.status_code == 200:
            data = response.json()
            
            if 'artists' in data and len(data['artists']) > 0:
                artist = data['artists'][0]
                
                # Extraer área de origen
                country = None
                city = None
                
                if 'area' in artist:
                    country = artist['area'].get('name')
                
                if 'begin-area' in artist:
                    city = artist['begin-area'].get('name')
                
                # Si tenemos al menos uno, devolver
                if country or city:
                    return {
                        'country': country,
                        'city': city
                    }
    except Exception as e:
        print(f"⚠️  Error obteniendo ubicación de {artist_name}: {str(e)}")
    
    return None


def get_instagram_stats(artist_name):
    """
    Obtiene estadísticas de Instagram del artista
    
    Args:
        artist_name (str): Nombre del artista
    
    Returns:
        dict: Username y seguidores, o None si hay error
    """
    # Si no hay RapidAPI key, retornar None
    if not RAPIDAPI_KEY or RAPIDAPI_KEY == '' or 'YOUR_RAPIDAPI_KEY' in RAPIDAPI_KEY:
        return None
    
    import requests
    
    try:
        # Simplificar nombre del artista para username
        username = artist_name.lower().replace(' ', '').replace('ñ', 'n')
        
        url = f"https://{INSTAGRAM_API_HOST}/api/instagram/user"
        
        querystring = {"username": username}
        
        headers = {
            "x-rapidapi-key": RAPIDAPI_KEY,
            "x-rapidapi-host": INSTAGRAM_API_HOST
        }
        
        response = requests.get(url, headers=headers, params=querystring, timeout=10)
        
        if response.status_code == 200:
            data = response.json()
            
            # Diferentes estructuras posibles
            followers = None
            
            if 'follower_count' in data:
                followers = data['follower_count']
            elif 'followers' in data:
                followers = data['followers']
            elif 'data' in data and 'follower_count' in data['data']:
                followers = data['data']['follower_count']
            
            if followers is not None:
                return {
                    'username': username,
                    'followers': followers
                }
        
    except Exception as e:
        print(f"⚠️  Error obteniendo Instagram de {artist_name}: {str(e)}")
    
    return None


def get_tiktok_stats(artist_name):
    """
    Obtiene estadísticas de TikTok del artista
    
    Args:
        artist_name (str): Nombre del artista
    
    Returns:
        dict: Username y seguidores, o None si hay error
    """
    # Si no hay RapidAPI key, retornar None
    if not RAPIDAPI_KEY or RAPIDAPI_KEY == '' or 'YOUR_RAPIDAPI_KEY' in RAPIDAPI_KEY:
        return None
    
    import requests
    
    try:
        # Simplificar nombre del artista para username
        username = artist_name.lower().replace(' ', '').replace('ñ', 'n')
        
        url = f"https://{TIKTOK_API_HOST}/api/user/info"
        
        querystring = {"username": username}
        
        headers = {
            "x-rapidapi-key": RAPIDAPI_KEY,
            "x-rapidapi-host": TIKTOK_API_HOST
        }
        
        response = requests.get(url, headers=headers, params=querystring, timeout=10)
        
        if response.status_code == 200:
            data = response.json()
            
            # Diferentes estructuras posibles
            followers = None
            
            if 'followerCount' in data:
                followers = data['followerCount']
            elif 'stats' in data and 'followerCount' in data['stats']:
                followers = data['stats']['followerCount']
            elif 'userInfo' in data and 'stats' in data['userInfo']:
                followers = data['userInfo']['stats'].get('followerCount')
            
            if followers is not None:
                return {
                    'username': username,
                    'followers': followers
                }
        
    except Exception as e:
        print(f"⚠️  Error obteniendo TikTok de {artist_name}: {str(e)}")
    
    return None


# ============================================
# RUTAS: Autenticación
# ============================================

@app.route('/register', methods=['GET', 'POST'])
def register():
    """
    Página de registro de nuevos usuarios
    GET: Muestra el formulario
    POST: Procesa el registro
    """
    if request.method == 'POST':
        data = request.get_json()
        
        username = data.get('username')
        email = data.get('email')
        password = data.get('password')
        
        # Validaciones básicas
        if not username or not email or not password:
            return jsonify({'error': 'Todos los campos son requeridos'}), 400
        
        if len(password) < 6:
            return jsonify({'error': 'Password debe tener al menos 6 caracteres'}), 400
        
        # Crear usuario
        user, error = create_user(username, email, password)
        
        if error:
            return jsonify({'error': error}), 400
        
        # Login automático después de registro
        login_user(user)
        
        return jsonify({
            'success': True,
            'message': 'Usuario creado exitosamente',
            'username': user.username
        })
    
    # GET request - mostrar formulario
    return render_template('register.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    """
    Página de login
    GET: Muestra el formulario
    POST: Procesa el login
    """
    if request.method == 'POST':
        data = request.get_json()
        
        email = data.get('email')
        password = data.get('password')
        
        if not email or not password:
            return jsonify({'error': 'Email y password requeridos'}), 400
        
        # Buscar usuario
        user = get_user_by_email(email)
        
        if not user:
            return jsonify({'error': 'Email o password incorrectos'}), 401
        
        # Verificar password
        if not verify_password(user, password):
            return jsonify({'error': 'Email o password incorrectos'}), 401
        
        # Login exitoso
        login_user(user)
        
        return jsonify({
            'success': True,
            'message': 'Login exitoso',
            'username': user.username
        })
    
    # GET request - mostrar formulario
    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    """
    Cerrar sesión
    @login_required significa que solo usuarios logueados pueden acceder
    """
    logout_user()
    return redirect(url_for('login'))


# ============================================
# RUTAS: Páginas principales
# ============================================

@app.route('/')
def home():
    """
    Página principal
    Si está logueado → Dashboard
    Si no está logueado → Página de bienvenida con login
    """
    if current_user.is_authenticated:
        # Usuario logueado, mostrar herramientas
        return render_template('index.html')
    else:
        # No logueado, redirigir a login
        return redirect(url_for('login'))


@app.route('/catalog')
@login_required  # NUEVO: Requiere login
def catalog():
    """
    Página para ver el catálogo
    AHORA: Solo muestra tracks del usuario actual
    """
    return render_template('catalog.html', is_admin=current_user.is_admin)


# ============================================
# RUTAS: Administración (solo para admins)
# ============================================

@app.route('/admin')
@login_required
def admin_panel():
    """
    Panel de administración
    Solo accesible para usuarios admin
    """
    if not current_user.is_admin:
        return redirect(url_for('home'))
    
    return render_template('admin.html')


@app.route('/api/admin/users', methods=['GET'])
@login_required
def admin_get_users():
    """
    API para obtener todos los usuarios (solo admin)
    """
    if not current_user.is_admin:
        return jsonify({'error': 'No autorizado'}), 403
    
    users = get_all_users()
    stats = get_global_stats()
    
    return jsonify({
        'users': users,
        'stats': stats
    })


@app.route('/api/admin/user/<int:user_id>/catalog', methods=['GET'])
@login_required
def admin_get_user_catalog(user_id):
    """
    API para ver el catálogo de un usuario específico (solo admin)
    """
    if not current_user.is_admin:
        return jsonify({'error': 'No autorizado'}), 403
    
    data = get_user_catalog_admin(user_id)
    
    if not data:
        return jsonify({'error': 'Usuario no encontrado'}), 404
    
    return jsonify(data)


@app.route('/api/catalog', methods=['GET'])
@login_required  # NUEVO: Requiere login
def get_catalog():
    """
    API para obtener el catálogo
    AHORA: Solo devuelve tracks del usuario actual
    """
    # Obtener parámetro de búsqueda
    search_query = request.args.get('q', '').strip()
    
    # NUEVO: Usar user_id del usuario actual
    user_id = current_user.id
    
    if search_query:
        tracks = search_tracks(search_query, user_id)
    else:
        tracks = get_all_tracks(user_id, limit=100)
    
    # Stats del usuario
    stats = get_catalog_stats(user_id)
    
    return jsonify({
        'tracks': [track.to_dict() for track in tracks],
        'total': len(tracks),
        'stats': stats,
        'username': current_user.username  # NUEVO
    })


# ============================================
# RUTAS: Extracción de metadata
# ============================================

@app.route('/extract-from-playlist', methods=['POST'])
@login_required  # NUEVO: Requiere login
def extract_from_playlist():
    """
    Extrae metadata de playlist de Spotify
    
    NUEVO: Modos de velocidad con caché de artistas
    - fetch_mode='fast': Solo Spotify (2-3 segundos por track) ⚡
    - fetch_mode='complete': Todos los datos (8-10 segundos por track) 🔍
    """
    try:
        data = request.get_json()
        playlist_url = data.get('playlist_url')
        fetch_mode = data.get('fetch_mode', 'fast')  # NUEVO: 'fast' o 'complete'
        
        if not playlist_url:
            return jsonify({'error': 'URL de playlist requerida'}), 400
        
        print(f"⚡ Modo de extracción: {fetch_mode.upper()}")
        
        # ============================================
        # MEJORADO: Extraer playlist ID
        # ============================================
        playlist_id = None
        
        if 'spotify.com/playlist/' in playlist_url:
            playlist_id = playlist_url.split('/playlist/')[-1].split('?')[0].split('/')[0]
        elif 'spotify:playlist:' in playlist_url:
            playlist_id = playlist_url.split('spotify:playlist:')[-1]
        else:
            playlist_id = playlist_url
        
        print(f"🎵 Extrayendo playlist ID: {playlist_id}")
        
        # Verificar playlist
        try:
            playlist_info = sp.playlist(playlist_id, fields='name,owner,public')
            print(f"✅ Playlist: {playlist_info['name']} by {playlist_info['owner']['display_name']}")
        except Exception as e:
            print(f"❌ Error: {str(e)}")
            return jsonify({
                'error': f'No se puede acceder a esta playlist. Verifica que sea pública.',
                'details': str(e)
            }), 404
        
        user_id = current_user.id
        
        # ============================================
        # Paginación ilimitada
        # ============================================
        all_track_items = []
        results = sp.playlist_tracks(playlist_id)
        all_track_items.extend(results['items'])
        
        while results['next']:
            results = sp.next(results)
            all_track_items.extend(results['items'])
        
        total = len(all_track_items)
        print(f"📊 Total tracks: {total}")
        
        # ============================================
        # OPTIMIZADO: Procesar tracks SECUENCIALMENTE (evita errores de contexto)
        # ============================================
        from models import get_or_create_artist_cache
        import time
        
        tracks = []
        saved_count = 0
        duplicate_count = 0
        start_time = time.time()
        
        for index, item in enumerate(all_track_items):
            if not item['track']:
                continue
            
            track = item['track']
            
            try:
                # Datos básicos (siempre)
                track_name = track['name']
                artists = ', '.join([a['name'] for a in track['artists']])
                first_artist_name = track['artists'][0]['name'] if track['artists'] else ''
                first_artist_id = track['artists'][0]['id'] if track['artists'] else None
                track_id = track['id']
                
                # Obtener álbum info
                album = sp.album(track['album']['id'])
                upc = album['external_ids'].get('upc', '')
                isrc = track['external_ids'].get('isrc', '')
                
                # Imagen del álbum
                album_image = None
                if album.get('images') and len(album['images']) > 0:
                    album_image = album['images'][1]['url'] if len(album['images']) > 1 else album['images'][0]['url']
                
                # Audio features (siempre, es rápido)
                audio_features = get_audio_features(track_id)
                
                # Track data base
                track_data = {
                    'track_name': track_name,
                    'artists': artists,
                    'album': track['album']['name'],
                    'isrc': isrc,
                    'upc': upc,
                    'spotify_url': track['external_urls']['spotify'],
                    'track_id': track_id,
                    'album_image_url': album_image
                }
                
                # Agregar audio features
                if audio_features:
                    track_data.update(audio_features)
                
                # ============================================
                # OPTIMIZADO: Usar caché de artista
                # ============================================
                if first_artist_id:
                    # Determinar si obtener datos completos
                    fetch_complete = (fetch_mode == 'complete')
                    
                    # Obtener del caché o crear
                    artist_data = get_or_create_artist_cache(
                        first_artist_id, 
                        first_artist_name,
                        sp,  # Pasar Spotify client
                        fetch_complete=fetch_complete,
                        get_location_func=get_artist_location if fetch_complete else None,
                        get_instagram_func=get_instagram_stats if fetch_complete else None,
                        get_tiktok_func=get_tiktok_stats if fetch_complete else None
                    )
                    
                    # Agregar datos de artista
                    track_data['artist_id'] = artist_data['artist_id']
                    track_data['artist_followers'] = artist_data['spotify_followers']
                    track_data['artist_popularity'] = artist_data['spotify_popularity']
                    track_data['artist_genres'] = artist_data['spotify_genres']
                    track_data['artist_image_url'] = artist_data['artist_image_url']
                    
                    # Datos opcionales (solo si fetch_complete)
                    if fetch_complete:
                        if artist_data.get('artist_country'):
                            track_data['artist_country'] = artist_data['artist_country']
                            track_data['artist_city'] = artist_data['artist_city']
                        if artist_data.get('instagram_username'):
                            track_data['instagram_username'] = artist_data['instagram_username']
                            track_data['instagram_followers'] = artist_data['instagram_followers']
                        if artist_data.get('tiktok_username'):
                            track_data['tiktok_username'] = artist_data['tiktok_username']
                            track_data['tiktok_followers'] = artist_data['tiktok_followers']
                
                # ============================================
                # Last.fm data (solo si fetch_complete)
                # ============================================
                if fetch_mode == 'complete':
                    lastfm_data = get_lastfm_data(track_name, first_artist_name)
                    if lastfm_data:
                        track_data['lastfm_tags'] = lastfm_data['tags']
                        track_data['lastfm_playcount'] = lastfm_data['playcount']
                        track_data['lastfm_listeners'] = lastfm_data['listeners']
                
                tracks.append(track_data)
                
                # Guardar en DB
                result = add_track(track_data, user_id)
                if result:
                    saved_count += 1
                else:
                    duplicate_count += 1
                
                # Log progreso cada 5 tracks
                if (index + 1) % 5 == 0:
                    elapsed = time.time() - start_time
                    avg_per_track = elapsed / (index + 1)
                    remaining = (total - (index + 1)) * avg_per_track
                    print(f"⏳ {index + 1}/{total} tracks | {elapsed:.1f}s elapsed | ~{remaining:.1f}s remaining")
                
            except Exception as e:
                print(f"❌ Error track {index + 1}: {str(e)}")
                continue
        
        elapsed_total = time.time() - start_time
        print(f"✅ Completado en {elapsed_total:.1f}s | {saved_count} guardados, {duplicate_count} duplicados")
        
        return jsonify({
            'success': True,
            'total_tracks': len(tracks),
            'tracks': tracks,
            'saved_to_db': saved_count,
            'duplicates': duplicate_count,
            'elapsed_time': round(elapsed_total, 1),
            'fetch_mode': fetch_mode
        })
        
    except Exception as e:
        print(f"❌ Error general: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/create-playlist-from-isrcs', methods=['POST'])
@login_required
def create_playlist_from_isrcs():
    """Busca tracks por ISRC y obtiene todos los datos"""
    try:
        data = request.get_json()
        isrcs = data.get('isrcs', [])
        
        if not isrcs:
            return jsonify({'error': 'Lista de ISRCs requerida'}), 400
        
        user_id = current_user.id
        
        track_uris = []
        found_tracks = []
        not_found = []
        
        for isrc in isrcs:
            try:
                # Buscar track por ISRC
                results = sp.search(q=f'isrc:{isrc}', type='track', limit=1)
                
                if results['tracks']['items']:
                    track = results['tracks']['items'][0]
                    track_uris.append(track['uri'])
                    track_id = track['id']
                    
                    # Obtener datos básicos
                    track_data = {
                        'isrc': isrc,
                        'track_name': track['name'],
                        'artists': ', '.join([a['name'] for a in track['artists']]),
                        'album': track['album']['name'],
                        'track_id': track_id,
                        'spotify_url': track['external_urls']['spotify'],
                        'upc': track['album'].get('external_ids', {}).get('upc')
                    }
                    
                    # Obtener imagen del álbum
                    if track['album']['images']:
                        track_data['album_image_url'] = track['album']['images'][0]['url']
                    
                    # Obtener audio features
                    try:
                        audio_features = sp.audio_features([track_id])[0]
                        if audio_features:
                            track_data.update({
                                'tempo': audio_features.get('tempo'),
                                'energy': audio_features.get('energy'),
                                'danceability': audio_features.get('danceability'),
                                'valence': audio_features.get('valence'),
                                'acousticness': audio_features.get('acousticness'),
                                'instrumentalness': audio_features.get('instrumentalness'),
                                'key': audio_features.get('key'),
                                'mode': audio_features.get('mode'),
                                'time_signature': audio_features.get('time_signature')
                            })
                    except:
                        pass
                    
                    # Obtener datos del artista
                    try:
                        artist_id = track['artists'][0]['id']
                        artist_info = sp.artist(artist_id)
                        
                        track_data.update({
                            'artist_id': artist_id,
                            'artist_followers': artist_info['followers']['total'],
                            'artist_popularity': artist_info['popularity'],
                            'artist_genres': ', '.join(artist_info['genres'][:5])
                        })
                        
                        # Imagen del artista
                        if artist_info['images']:
                            track_data['artist_image_url'] = artist_info['images'][0]['url']
                    except:
                        pass
                    
                    # Obtener datos de Last.fm
                    try:
                        artist_name = track['artists'][0]['name']
                        lastfm_data = get_lastfm_data(artist_name, track['name'])
                        if lastfm_data:
                            track_data.update({
                                'lastfm_tags': lastfm_data.get('tags'),
                                'lastfm_playcount': lastfm_data.get('playcount'),
                                'lastfm_listeners': lastfm_data.get('listeners')
                            })
                    except:
                        pass
                    
                    # Obtener ubicación del artista (MusicBrainz)
                    try:
                        first_artist_name = track['artists'][0]['name']
                        location_data = get_artist_location(first_artist_name)
                        if location_data:
                            track_data['artist_country'] = location_data['country']
                            track_data['artist_city'] = location_data['city']
                    except:
                        pass
                    
                    # Obtener Instagram
                    try:
                        first_artist_name = track['artists'][0]['name']
                        instagram_data = get_instagram_stats(first_artist_name)
                        if instagram_data:
                            track_data['instagram_username'] = instagram_data['username']
                            track_data['instagram_followers'] = instagram_data['followers']
                    except:
                        pass
                    
                    # Obtener TikTok
                    try:
                        first_artist_name = track['artists'][0]['name']
                        tiktok_data = get_tiktok_stats(first_artist_name)
                        if tiktok_data:
                            track_data['tiktok_username'] = tiktok_data['username']
                            track_data['tiktok_followers'] = tiktok_data['followers']
                    except:
                        pass
                    
                    found_tracks.append(track_data)
                    
                    # Guardar con user_id
                    add_track(track_data, user_id)
                    
                    print(f"✅ Track importado por ISRC: {track_data['track_name']}")
                else:
                    not_found.append(isrc)
                    print(f"⚠️  ISRC no encontrado: {isrc}")
            except Exception as e:
                not_found.append(isrc)
                print(f"❌ Error con ISRC {isrc}: {str(e)}")
        
        return jsonify({
            'success': True,
            'found': len(found_tracks),
            'not_found': len(not_found),
            'tracks': found_tracks,
            'track_uris': track_uris,
            'missing_isrcs': not_found
        })
        
    except Exception as e:
        print(f"❌ Error en create_playlist_from_isrcs: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/download-excel', methods=['POST'])
@login_required
def download_excel():
    """Descarga tracks como Excel con columnas específicas"""
    try:
        data = request.get_json()
        tracks = data.get('tracks', [])
        
        if not tracks:
            return jsonify({'error': 'No hay tracks para descargar'}), 400
        
        # Filtrar solo las columnas que queremos
        filtered_tracks = []
        for track in tracks:
            filtered_tracks.append({
                'artists': track.get('artists', ''),
                'album': track.get('album', ''),
                'track_name': track.get('track_name', ''),
                'upc': track.get('upc', ''),
                'isrc': track.get('isrc', '')
            })
        
        df = pd.DataFrame(filtered_tracks)
        
        # Renombrar columnas para que se vean mejor
        df.columns = ['ARTIST', 'ALBUM NAME', 'TRACK NAME', 'UPC', 'ISRC']
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Tracks')
        output.seek(0)
        
        filename = f'songxs_{current_user.username}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================
# RUTAS: Export Catalog (PDF y Excel)
# ============================================

@app.route('/api/export-excel')
@login_required
def export_excel():
    """Export catalog to Excel"""
    try:
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        tracks = Track.query.filter_by(user_id=current_user.id).all()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Catalog"
        
        headers = ['ARTIST', 'ALBUM', 'TRACK', 'UPC', 'ISRC', 'CLEARANCE', 'PRICE']
        ws.append(headers)
        
        header_fill = PatternFill(start_color="1F1F1F", end_color="1F1F1F", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for track in tracks:
            ws.append([
                track.artists or '',
                track.album or '',
                track.track_name or '',
                track.upc or '',
                track.isrc or '',
                track.clearance_type or '',
                f'${track.clearance_price:.2f}' if track.clearance_price else ''
            ])
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'catalog_{current_user.username}.xlsx'
        )
    except Exception as e:
        print(f"❌ Error export-excel: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/export-pdf')
@login_required
def export_pdf():
    """Export catalog to PDF with logo and optimized layout"""
    try:
        from io import BytesIO
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
        import os
        
        tracks = Track.query.filter_by(user_id=current_user.id).all()
        
        buffer = BytesIO()
        # Use A4 landscape for more space
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=landscape(A4),
            topMargin=0.8*inch,
            bottomMargin=0.5*inch,
            leftMargin=0.5*inch,
            rightMargin=0.5*inch
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Add logo if exists
        logo_path = os.path.join('static', 'logo.png')
        if os.path.exists(logo_path):
            try:
                logo = Image(logo_path, width=1.5*inch, height=0.5*inch)
                elements.append(logo)
                elements.append(Spacer(1, 0.2*inch))
            except:
                pass  # Skip logo if error
        
        # Title
        title = Paragraph(f"<b>Music Catalog - {current_user.username}</b>", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 0.3*inch))
        
        # Prepare data
        data = [['ARTIST', 'ALBUM', 'TRACK', 'UPC', 'ISRC', 'CLEARANCE', 'PRICE']]
        
        for track in tracks:
            # Truncate long text to prevent overflow
            artist = (track.artists or '')[:30]
            album = (track.album or '')[:25]
            track_name = (track.track_name or '')[:30]
            
            data.append([
                artist,
                album,
                track_name,
                track.upc or '',
                track.isrc or '',
                track.clearance_type or '',
                f'${track.clearance_price:.2f}' if track.clearance_price else ''
            ])
        
        # Define column widths (A4 landscape = ~11 inches usable)
        # Total: ~10.5 inches for content
        col_widths = [
            1.5*inch,  # ARTIST
            1.5*inch,  # ALBUM
            1.8*inch,  # TRACK
            1.3*inch,  # UPC
            1.3*inch,  # ISRC
            1.0*inch,  # CLEARANCE
            0.8*inch   # PRICE
        ]
        
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.black),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            # Body
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('WORDWRAP', (0, 1), (-1, -1), True),  # Enable word wrap
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ]))
        
        elements.append(table)
        doc.build(elements)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'catalog_{current_user.username}.pdf'
        )
    except Exception as e:
        print(f"❌ Error export-pdf: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ============================================
# RUTAS: Clearance Management
# ============================================

@app.route('/api/track/<int:track_id>/clearance', methods=['PUT'])
@login_required
def update_clearance(track_id):
    """Update track clearance type and price"""
    try:
        data = request.get_json()
        
        track = Track.query.filter_by(id=track_id, user_id=current_user.id).first()
        
        if not track:
            return jsonify({'error': 'Track not found'}), 404
        
        clearance_type = data.get('clearance_type')
        if clearance_type not in ['EasyClear', 'PreClear', 'FreeClear', None]:
            return jsonify({'error': 'Invalid clearance type'}), 400
        
        track.clearance_type = clearance_type
        
        if clearance_type == 'FreeClear':
            track.clearance_price = 0
        else:
            price = data.get('clearance_price')
            if price is not None:
                try:
                    track.clearance_price = float(price)
                except:
                    return jsonify({'error': 'Invalid price'}), 400
        
        db.session.commit()
        
        print(f"✅ Clearance updated: Track {track_id} → {clearance_type} @ ${track.clearance_price or 0}")
        
        return jsonify({
            'success': True,
            'clearance_type': track.clearance_type,
            'clearance_price': track.clearance_price
        })
    except Exception as e:
        print(f"❌ Error update_clearance: {e}")
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ============================================
# RUTAS: Split Sheets
# ============================================

@app.route('/api/track/<int:track_id>/split-sheet', methods=['POST'])
@login_required
def create_split_sheet_route(track_id):
    """Crea un nuevo split sheet para un track (SOLO ADMIN)"""
    try:
        # VERIFICAR QUE ES ADMIN
        if not current_user.is_admin:
            return jsonify({'error': 'Only admins can create split sheets'}), 403
        
        from models import create_split_sheet
        
        data = request.get_json()
        
        # Verificar que el track pertenece al usuario
        track = Track.query.filter_by(id=track_id, user_id=current_user.id).first()
        if not track:
            return jsonify({'error': 'Track not found'}), 404
        
        collaborators_data = data.get('collaborators', [])
        notes = data.get('notes', '')
        
        if not collaborators_data:
            return jsonify({'error': 'At least one collaborator is required'}), 400
        
        # Validar que suma 100%
        total_percentage = sum(c.get('percentage', 0) for c in collaborators_data)
        if abs(total_percentage - 100) > 0.01:
            return jsonify({'error': f'Percentages must sum to 100% (currently {total_percentage}%)'}), 400
        
        # Crear split sheet
        split_sheet = create_split_sheet(track_id, current_user.id, collaborators_data, notes)
        
        print(f"✅ Split sheet created: {split_sheet.id} for track {track_id}")
        
        return jsonify({
            'success': True,
            'split_sheet': split_sheet.to_dict()
        })
    except Exception as e:
        print(f"❌ Error creating split sheet: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/track/<int:track_id>/split-sheets', methods=['GET'])
@login_required
def get_track_split_sheets(track_id):
    """Obtiene todos los split sheets de un track (SOLO ADMIN)"""
    try:
        # VERIFICAR QUE ES ADMIN
        if not current_user.is_admin:
            return jsonify({'error': 'Only admins can view split sheets'}), 403
        
        from models import get_split_sheets_by_track
        
        # Verificar que el track pertenece al usuario
        track = Track.query.filter_by(id=track_id, user_id=current_user.id).first()
        if not track:
            return jsonify({'error': 'Track not found'}), 404
        
        split_sheets = get_split_sheets_by_track(track_id)
        
        return jsonify({
            'success': True,
            'split_sheets': [ss.to_dict() for ss in split_sheets]
        })
    except Exception as e:
        print(f"❌ Error getting split sheets: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/split-sheet/<int:split_sheet_id>', methods=['GET'])
@login_required
def get_split_sheet_route(split_sheet_id):
    """Obtiene un split sheet específico (SOLO ADMIN)"""
    try:
        # VERIFICAR QUE ES ADMIN
        if not current_user.is_admin:
            return jsonify({'error': 'Only admins can view split sheets'}), 403
        
        from models import get_split_sheet
        
        split_sheet = get_split_sheet(split_sheet_id)
        if not split_sheet:
            return jsonify({'error': 'Split sheet not found'}), 404
        
        # Verificar que el usuario tiene acceso
        if split_sheet.created_by_user_id != current_user.id:
            return jsonify({'error': 'Access denied'}), 403
        
        return jsonify({
            'success': True,
            'split_sheet': split_sheet.to_dict()
        })
    except Exception as e:
        print(f"❌ Error getting split sheet: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/sign/<signature_token>')
def sign_page(signature_token):
    """Página de firma para colaboradores"""
    from models import Collaborator
    
    collaborator = Collaborator.query.filter_by(signature_token=signature_token).first()
    
    if not collaborator:
        return render_template('error.html', 
            error='Invalid signature link',
            message='This signature link is not valid or has expired.'
        ), 404
    
    if collaborator.signed:
        return render_template('already_signed.html', 
            collaborator=collaborator
        )
    
    split_sheet = collaborator.split_sheet
    track = split_sheet.track
    
    return render_template('sign_split_sheet.html',
        collaborator=collaborator,
        split_sheet=split_sheet,
        track=track
    )


@app.route('/api/sign/<signature_token>', methods=['POST'])
def sign_collaborator_route(signature_token):
    """Procesa la firma de un colaborador"""
    try:
        from models import sign_collaborator
        
        data = request.get_json()
        signature_data = data.get('signature')
        
        if not signature_data:
            return jsonify({'error': 'Signature data required'}), 400
        
        success = sign_collaborator(signature_token, signature_data)
        
        if not success:
            return jsonify({'error': 'Invalid signature token'}), 404
        
        print(f"✅ Collaborator signed with token {signature_token[:10]}...")
        
        return jsonify({
            'success': True,
            'message': 'Signature recorded successfully'
        })
    except Exception as e:
        print(f"❌ Error signing: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/split-sheet/<int:split_sheet_id>/pdf')
@login_required
def download_split_sheet_pdf(split_sheet_id):
    """Genera y descarga el PDF del split sheet (SOLO ADMIN)"""
    try:
        # VERIFICAR QUE ES ADMIN
        if not current_user.is_admin:
            return jsonify({'error': 'Only admins can download split sheet PDFs'}), 403
        
        from models import get_split_sheet
        from io import BytesIO
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        import os
        
        split_sheet = get_split_sheet(split_sheet_id)
        if not split_sheet:
            return jsonify({'error': 'Split sheet not found'}), 404
        
        # Verificar acceso
        if split_sheet.created_by_user_id != current_user.id:
            return jsonify({'error': 'Access denied'}), 403
        
        track = split_sheet.track
        
        # Crear PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            topMargin=0.75*inch,
            bottomMargin=0.75*inch,
            leftMargin=0.75*inch,
            rightMargin=0.75*inch
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Logo
        logo_path = os.path.join('static', 'logo.png')
        if os.path.exists(logo_path):
            try:
                logo = RLImage(logo_path, width=1.5*inch, height=0.5*inch)
                elements.append(logo)
                elements.append(Spacer(1, 0.3*inch))
            except:
                pass
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.black,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        title = Paragraph("SPLIT SHEET AGREEMENT", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.3*inch))
        
        # Información del track
        track_info_style = ParagraphStyle(
            'TrackInfo',
            parent=styles['Normal'],
            fontSize=11,
            spaceAfter=6
        )
        
        elements.append(Paragraph(f"<b>Track:</b> {track.track_name}", track_info_style))
        elements.append(Paragraph(f"<b>Artist(s):</b> {track.artists}", track_info_style))
        elements.append(Paragraph(f"<b>Album:</b> {track.album or 'N/A'}", track_info_style))
        elements.append(Paragraph(f"<b>ISRC:</b> {track.isrc or 'N/A'}", track_info_style))
        elements.append(Spacer(1, 0.3*inch))
        
        # Tabla de colaboradores
        collab_data = [['Name', 'Role', 'Percentage', 'IPI/CAE', 'PRO', 'Signed']]
        
        for collab in split_sheet.collaborators:
            collab_data.append([
                collab.name,
                collab.role or 'N/A',
                f"{collab.percentage}%",
                collab.ipi_cae or 'N/A',
                collab.pro or 'N/A',
                '✓' if collab.signed else '✗'
            ])
        
        # Agregar fila de total
        collab_data.append([
            'TOTAL',
            '',
            f"{split_sheet.total_percentage()}%",
            '',
            '',
            ''
        ])
        
        table = Table(collab_data, colWidths=[1.8*inch, 1.0*inch, 0.9*inch, 1.0*inch, 1.0*inch, 0.7*inch])
        table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.black),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            # Body
            ('BACKGROUND', (0, 1), (-1, -2), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            # Total row
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 0.4*inch))
        
        # Notas
        if split_sheet.notes:
            elements.append(Paragraph("<b>Notes:</b>", styles['Normal']))
            elements.append(Paragraph(split_sheet.notes, styles['Normal']))
            elements.append(Spacer(1, 0.3*inch))
        
        # Firmas (si hay)
        if split_sheet.all_signed():
            elements.append(Paragraph("<b>SIGNATURES</b>", styles['Heading2']))
            elements.append(Spacer(1, 0.2*inch))
            
            for collab in split_sheet.collaborators:
                if collab.signed and collab.signature_data:
                    elements.append(Paragraph(f"<b>{collab.name}</b>", styles['Normal']))
                    elements.append(Paragraph(f"Signed on: {collab.signed_at.strftime('%Y-%m-%d %H:%M UTC')}", styles['Normal']))
                    elements.append(Spacer(1, 0.1*inch))
        
        # Footer
        elements.append(Spacer(1, 0.5*inch))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        elements.append(Paragraph(
            f"Generated on {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} | Status: {split_sheet.status}",
            footer_style
        ))
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"split_sheet_{track.track_name.replace(' ', '_')}_{split_sheet_id}.pdf"
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print(f"❌ Error generating PDF: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ============================================
# EJECUTAR APP
# ============================================
if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
